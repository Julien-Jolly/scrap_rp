# OK - mettre tous les fichiers du repertoire courant dans une liste
# OK - récupérer la 1ere date en A3 dans une variable
# OK - récupérer les data de chaque feuille à partir de la ligne 8 et jusqu'à 1ere ligne vide
# select colonnes à récupérer :
# Competition : A-H / Categories Negatively Affecting : A-F / Indexes Evolution : B-H
# ajouter la date dans une nouvelle colonne à droite
# concatener les données des fichiers dans une liste


import openpyxl
import os
from datetime import datetime
import csv
import main


def check_rows(file, sheets):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheets[index]]
    count = 0

    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not any(row):
            break
        else:
            count += 1

    return count


def list_files(directory):
    files = []
    for filename in directory:
        if os.path.isfile(os.path.join(directory, filename)):
            files.append(directory + "/" + filename)
    return files


def list_dir(mdirectory):
    dir = []
    for filename in os.listdir(mdirectory):
        dir.append(mdirectory + filename)
    return dir


def extract_date(file_path, sheet_name, cellule_date):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell_value = sheet[cellule_date].value
    date = cell_value[12:23]
    original_date = datetime.strptime(date, "%d %b %Y")
    formated_date = original_date.strftime("%d/%m/%Y")

    workbook.close()

    return formated_date


def extract_note(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Booking.com":
                adjacent_cell_value = sheet.cell(
                    row=cell.row, column=cell.column + 1
                ).value
                return adjacent_cell_value


def sheets_names(file):
    workbook = openpyxl.load_workbook(file)
    sheet_names = workbook.sheetnames
    return sheet_names


def concat(PATH, option_path, year, mois_str):

    mdirectory = f"{PATH}{option_path}{mois_str}-{year}/"
    sheet = "Source Profile Information"
    columns = ["A:F", "A:F"]
    cellule_date = "A3"
    # choix = int(input("Quelle feuille ? :"))
    choix = 1
    index = choix - 1
    headers = ["Hotel", "Date", "Note"]
    extract = []

    hotels_id=main.hotels_id

    dir_list = list_dir(mdirectory)
    print(f"contenu du repertoire {mdirectory}: {dir_list}")

    for i in range(len(dir_list)):

        print(f"fichier traité : {dir_list[i]}")

        formated_date = extract_date(dir_list[i], sheet, cellule_date)
        note = extract_note(dir_list[i], sheet)
        if note:
            line = hotels_id[i], formated_date, note[:3]
        else:
            line = hotels_id[i], formated_date, "NA"
        extract.append(line)
        print(f"extraction : {extract}")

    csv_file_path = f"{PATH}{option_path}booking_{mois_str}-{year}.csv"

    # Écriture des données dans le fichier CSV
    with open(
        csv_file_path,
        mode="w",
        newline="",
    ) as file:
        writer = csv.writer(file)
        writer.writerow(headers)
        writer.writerows(extract)

    print(f"Les données ont été enregistrées dans le fichier {csv_file_path}")
