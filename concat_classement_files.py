import openpyxl
import os
from datetime import datetime
import csv
import main

def extract_data_indexes_evolution(file_path, sheet_name, mois_str):
    """Extrait les données de la feuille Indexes Evolution pour le mois spécifié."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=False)
        sheet = workbook[sheet_name]

        data = []
        print(f"Feuille {sheet_name} ouverte dans {file_path}")
        print(f"Nombre total de lignes dans la feuille: {sheet.max_row}")

        for i, row in enumerate(sheet.rows, start=1):
            row_values = [cell.value for cell in row]
            if i < 8:
                print(f"Ligne {i}, contenu brut (ignoré): {row_values}")
                continue
            print(f"Ligne {i}, contenu brut: {row_values}")
            if not any(cell.value for cell in row):
                print(f"Ligne {i} vide, arrêt de l'extraction")
                break

            date_str = row_values[0] if row_values else None
            print(f"Ligne {i}, date brute: {date_str}, type: {type(date_str) if date_str else 'NoneType'}")

            if date_str:
                try:
                    if isinstance(date_str, datetime):
                        row_month = f"{date_str.month:02d}"
                        date_formatted = date_str.strftime("%Y-%m-%d")
                        print(f"Date convertie depuis datetime: {date_formatted}, mois: {row_month}")
                    else:
                        date_str = str(date_str).strip()
                        if '-' in date_str:
                            row_month = date_str.split('-')[1]
                            print(f"Mois extrait de la chaîne: {row_month}")
                        else:
                            print(f"Format de date non reconnu: {date_str}")
                            continue

                    row_month = row_month.zfill(2)
                    print(f"Mois formaté: {row_month}, mois_str: {mois_str}")

                    if row_month == mois_str:
                        row_data = row_values[:8]
                        if isinstance(date_str, datetime):
                            row_data[0] = date_str.strftime("%Y-%m-%d")
                        data.append(row_data)
                        print(f"Ligne ajoutée: {row_data}")
                    else:
                        print(f"Mois {row_month} ne correspond pas à mois_str {mois_str}")
                except (IndexError, ValueError, AttributeError) as e:
                    print(f"Erreur de parsing de la date à la ligne {i}: {e}")
                    continue
            else:
                print(f"Date vide à la ligne {i}")

        workbook.close()
        print(f"data : {data}")
        return data
    except Exception as e:
        print(f"Erreur lors de l'extraction des données dans {file_path}: {e}")
        return []

def extract_data_competition(file_path, sheet_name, mois_str, year, hotel_id):
    """Extrait les données de la feuille Competition pour le mois spécifié avec décalage pour hotel_id."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=False)
        sheet = workbook[sheet_name]

        data = []
        print(f"Feuille {sheet_name} ouverte dans {file_path}")
        print(f"Nombre total de lignes dans la feuille: {sheet.max_row}")

        for i, row in enumerate(sheet.rows, start=1):
            row_values = [cell.value for cell in row]
            if i < 8:
                print(f"Ligne {i}, contenu brut (ignoré): {row_values}")
                continue
            print(f"Ligne {i}, contenu brut: {row_values}")
            if not any(cell.value for cell in row):
                print(f"Ligne {i} vide, arrêt de l'extraction")
                break

            hotel_name = row_values[0] if row_values else None
            print(f"Ligne {i}, hotel brute: {hotel_name}, type: {type(hotel_name) if hotel_name else 'NoneType'}")

            if hotel_name:
                # Conserver toutes les 6 colonnes originales et ajouter hotel_id en première position, puis Mois/Année en dernière
                # Formater Mois/Année en jj/mm/aaaa (premier jour du mois)
                mois_annee = f"01/{mois_str}/{year}"
                row_data = [hotel_id] + row_values[:6] + [mois_annee]
                data.append(row_data)
                print(f"Ligne ajoutée: {row_data}")
            else:
                print(f"Hotel vide à la ligne {i}")

        workbook.close()
        print(f"data : {data}")
        return data
    except Exception as e:
        print(f"Erreur lors de l'extraction des données dans {file_path}: {e}")
        return []

def extract_data_categories_negatively_affecting(file_path, sheet_name, mois_str, year, hotel_id):
    """Extrait les données de la feuille Categories Negatively Affecting pour le mois spécifié avec décalage pour hotel_id."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=False)
        sheet = workbook[sheet_name]

        data = []
        print(f"Feuille {sheet_name} ouverte dans {file_path}")
        print(f"Nombre total de lignes dans la feuille: {sheet.max_row}")

        for i, row in enumerate(sheet.rows, start=1):
            row_values = [cell.value for cell in row]
            if i < 8:
                print(f"Ligne {i}, contenu brut (ignoré): {row_values}")
                continue
            print(f"Ligne {i}, contenu brut: {row_values}")
            if not any(cell.value for cell in row):
                print(f"Ligne {i} vide, arrêt de l'extraction")
                break

            category_name = row_values[0] if row_values else None
            print(
                f"Ligne {i}, catégorie brute: {category_name}, type: {type(category_name) if category_name else 'NoneType'}")

            if category_name:
                # Conserver toutes les 6 colonnes originales et ajouter hotel_id en première position, puis Mois/Année en dernière
                # Formater Mois/Année en jj/mm/aaaa (premier jour du mois)
                mois_annee = f"01/{mois_str}/{year}"
                row_data = [hotel_id] + row_values[:6] + [mois_annee]
                data.append(row_data)
                print(f"Ligne ajoutée: {row_data}")
            else:
                print(f"Catégorie vide à la ligne {i}")

        workbook.close()
        print(f"data : {data}")
        return data
    except Exception as e:
        print(f"Erreur lors de l'extraction des données dans {file_path}: {e}")
        return []

def list_dir(mdirectory):
    """Liste les fichiers dans un répertoire."""
    try:
        return [os.path.join(mdirectory, f) for f in os.listdir(mdirectory) if
                os.path.isfile(os.path.join(mdirectory, f))]
    except Exception as e:
        print(f"Erreur lors de la lecture du répertoire {mdirectory}: {e}")
        return []

def format_value(value):
    """Formate les valeurs pour correspondre au format attendu (virgules pour décimaux)."""
    if value is None:
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:.1f}".replace(".", ",")
    if isinstance(value, str):
        if "%" in value:
            return value.replace(".", ",")
        # Remplacer les points par des virgules pour les nombres décimaux
        try:
            return str(float(value)).replace(".", ",")
        except ValueError:
            return value
    return str(value)

def parse_and_format_date(date_str):
    """Parse et reformate une date en jj/mm/aaaa."""
    if not date_str or not isinstance(date_str, str):
        return "01/01/1900"  # Date par défaut si invalide
    try:
        # Tenter différents formats possibles
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "mmm-yy"):
            try:
                if fmt == "mmm-yy":
                    # Convertir "avr-25" en "01/04/2025"
                    month_abbr = date_str.split('-')[0].lower()
                    year = "20" + date_str.split('-')[1]
                    month_map = {
                        "jan": "01", "fév": "02", "mar": "03", "avr": "04",
                        "mai": "05", "jun": "06", "jul": "07", "aoû": "08",
                        "sep": "09", "oct": "10", "nov": "11", "déc": "12"
                    }
                    month = month_map.get(month_abbr, "01")
                    return f"01/{month}/{year}"
                else:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime("%d/%m/%Y")
            except ValueError:
                continue
        print(f"Format de date non reconnu, valeur brute : {date_str}")
        return "01/01/1900"  # Date par défaut si aucun format ne correspond
    except Exception as e:
        print(f"Erreur lors du parsing de la date {date_str}: {e}")
        return "01/01/1900"

def concat(PATH, option_path, year, mois_str):
    """Concatène les données des fichiers téléchargés pour le choix 1 dans des fichiers individuels par hôtel,
    puis regroupe tout dans un fichier unique par typologie."""
    mdirectory = f"{PATH}{option_path}{mois_str}-{year}/"
    consolidated_dir_indexes = os.path.join(PATH, "Fait", "Index evolution")
    consolidated_dir_competition = os.path.join(PATH, "Fait", "Competition")
    consolidated_dir_categories = os.path.join(PATH, "Fait", "Categories")

    # Créer les dossiers consolidés s'ils n'existent pas
    for dir_path in [consolidated_dir_indexes, consolidated_dir_competition, consolidated_dir_categories]:
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
            print(f"Création du répertoire consolidé : {dir_path}")

    headers_indexes = ["Hotel_Ref", "Date", "GRI", "Change", "Goal", "Reviews", "Change", "Mentions", "Change"]
    headers_competition = ["Hotel_Ref", "Hotel", "Index", "Change", "Reviews", "Change", "CQI™", "Mois/Année"]
    headers_categories = ["Hotel_Ref", "Categories", "Negative Mentions", "Change", "GRI Impact", "Change",
                          "Top Concept", "Mois/Année"]

    # Liste des fichiers dans le répertoire
    file_list = list_dir(mdirectory)
    print(f"Contenu du répertoire {mdirectory}: {file_list}")

    # Utilisation de hotels_data depuis main.py (si défini)
    hotels_data = getattr(main, 'hotels_data', None)
    if hotels_data is None:
        hotels_id = main.hotels_id
        hotels = main.hotels
        hotels_data = [{"id": hotels_id[i], "name": hotels[i]} for i in range(len(hotels))]

    # Étape 1 : Concaténation dans des fichiers individuels par hôtel
    for file_path in file_list:
        try:
            hotel_name = os.path.splitext(os.path.basename(file_path))[0]
            hotel_id = next((h["id"] for h in hotels_data if h["name"] == hotel_name), "Inconnu")

            print(f"Fichier traité : {file_path} (Hotel: {hotel_name}, ID: {hotel_id})")

            # Extraction des données pour Indexes Evolution
            data_indexes = extract_data_indexes_evolution(file_path, "Indexes Evolution", mois_str)
            formatted_data_indexes = []
            for row in data_indexes:
                formatted_row = [hotel_id] + [format_value(cell) for cell in row]
                if formatted_row[1] and isinstance(formatted_row[1], str):
                    try:
                        date_obj = datetime.strptime(formatted_row[1], "%Y-%m-%d")
                        formatted_row[1] = date_obj.strftime("%d/%m/%Y")
                    except ValueError as e:
                        print(f"Erreur lors du formatage de la date {formatted_row[1]}: {e}")
                formatted_data_indexes.append(formatted_row)

            consolidated_file_indexes = os.path.join(consolidated_dir_indexes, f"{hotel_id}.csv")
            file_exists_indexes = os.path.exists(consolidated_file_indexes)
            with open(consolidated_file_indexes, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=";")
                if not file_exists_indexes:
                    writer.writerow(headers_indexes)
                writer.writerows(formatted_data_indexes)

            print(f"Données Indexes ajoutées au fichier individuel : {consolidated_file_indexes}")

            # Extraction des données pour Competition
            data_competition = extract_data_competition(file_path, "Competition", mois_str, year, hotel_id)
            formatted_data_competition = []
            for row in data_competition:
                formatted_row = [format_value(cell) for cell in row]
                formatted_data_competition.append(formatted_row)

            consolidated_file_competition = os.path.join(consolidated_dir_competition, f"{hotel_id}.csv")
            file_exists_competition = os.path.exists(consolidated_file_competition)
            with open(consolidated_file_competition, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=";")
                if not file_exists_competition:
                    writer.writerow(headers_competition)
                writer.writerows(formatted_data_competition)

            print(f"Données Competition ajoutées au fichier individuel : {consolidated_file_competition}")

            # Extraction des données pour Categories Negatively Affecting
            data_categories = extract_data_categories_negatively_affecting(file_path, "Categories Negatively Affecting",
                                                                           mois_str, year, hotel_id)
            formatted_data_categories = []
            for row in data_categories:
                formatted_row = [format_value(cell) for cell in row]
                formatted_data_categories.append(formatted_row)

            consolidated_file_categories = os.path.join(consolidated_dir_categories, f"{hotel_id}.csv")
            file_exists_categories = os.path.exists(consolidated_file_categories)
            with open(consolidated_file_categories, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=";")
                if not file_exists_categories:
                    writer.writerow(headers_categories)
                writer.writerows(formatted_data_categories)

            print(f"Données Categories ajoutées au fichier individuel : {consolidated_file_categories}")

            print(
                f"Données extraites pour {hotel_name}: {len(data_indexes)} lignes (Indexes), {len(data_competition)} lignes (Competition), {len(data_categories)} lignes (Categories)")

        except Exception as e:
            print(f"Erreur lors du traitement du fichier {file_path}: {e}")

    # Étape 2 : Regroupement dans des fichiers uniques pour chaque typologie
    # Index Evolution
    consolidated_unique_indexes = os.path.join(consolidated_dir_indexes, "index_evolution.csv")
    all_data_indexes = []
    for file_path in list_dir(consolidated_dir_indexes):
        if os.path.basename(file_path) != "index_evolution.csv":  # Ignorer le fichier consolidé lui-même
            try:
                with open(file_path, mode="r", newline="", encoding="utf-8", errors="replace") as file:
                    reader = csv.reader(file, delimiter=";")
                    next(reader)  # Ignorer l'en-tête du fichier individuel
                    for row in reader:
                        # Reformater la date (colonne 1 pour Indexes)
                        if row[1]:
                            row[1] = parse_and_format_date(row[1])
                        all_data_indexes.append(row)
            except Exception as e:
                print(f"Erreur lors de la lecture du fichier {file_path} pour Index Evolution: {e}")

    with open(consolidated_unique_indexes, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(headers_indexes)
        writer.writerows(all_data_indexes)

    print(f"Toutes les données Indexes regroupées dans : {consolidated_unique_indexes}")

    # Competition
    consolidated_unique_competition = os.path.join(consolidated_dir_competition, "competition.csv")
    all_data_competition = []
    for file_path in list_dir(consolidated_dir_competition):
        if os.path.basename(file_path) != "competition.csv":  # Ignorer le fichier consolidé lui-même
            try:
                with open(file_path, mode="r", newline="", encoding="utf-8", errors="replace") as file:
                    reader = csv.reader(file, delimiter=";")
                    next(reader)  # Ignorer l'en-tête du fichier individuel
                    for row in reader:
                        # Reformater la date (colonne 7 pour Competition)
                        if row[7]:
                            row[7] = parse_and_format_date(row[7])
                        all_data_competition.append(row)
            except Exception as e:
                print(f"Erreur lors de la lecture du fichier {file_path} pour Competition: {e}")

    with open(consolidated_unique_competition, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(headers_competition)
        writer.writerows(all_data_competition)

    print(f"Toutes les données Competition regroupées dans : {consolidated_unique_competition}")

    # Categories
    consolidated_unique_categories = os.path.join(consolidated_dir_categories, "categories.csv")
    all_data_categories = []
    for file_path in list_dir(consolidated_dir_categories):
        if os.path.basename(file_path) != "categories.csv":  # Ignorer le fichier consolidé lui-même
            try:
                with open(file_path, mode="r", newline="", encoding="utf-8", errors="replace") as file:
                    reader = csv.reader(file, delimiter=";")
                    next(reader)  # Ignorer l'en-tête du fichier individuel
                    for row in reader:
                        # Reformater la date (colonne 7 pour Categories)
                        if row[7]:
                            row[7] = parse_and_format_date(row[7])
                        all_data_categories.append(row)
            except Exception as e:
                print(f"Erreur lors de la lecture du fichier {file_path} pour Categories: {e}")

    with open(consolidated_unique_categories, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(headers_categories)
        writer.writerows(all_data_categories)

    print(f"Toutes les données Categories regroupées dans : {consolidated_unique_categories}")